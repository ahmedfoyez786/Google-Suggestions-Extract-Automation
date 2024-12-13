from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
import random
import time

# Get the current day
x = datetime.datetime.now()
today = x.strftime("%A")

# Load Excel file
file_path = "Excel.xlsx"
workbook = load_workbook(file_path)

if today in workbook.sheetnames:
    sheet = workbook[today]
else:
    print(f"No sheet found for {today}")
    exit()

# Read keywords from column C
data = [cell.value for cell in sheet["C"] if cell.value not in (None, "C")]

# Set up WebDriver
driver_path = r"C:\\Development\\msedgedriver.exe"
edge_service = Service(executable_path=driver_path)
edge_option = webdriver.EdgeOptions()
edge_option.add_experimental_option("detach", True)

driver = webdriver.Edge(service=edge_service, options=edge_option)
driver.get("https://www.google.com/")
time.sleep(random.uniform(1.0, 2.0))

# Perform Google searches and store suggestions
def fetch_suggestions(keyword):
    search_box = driver.find_element(By.XPATH, '//*[@title="Search"]')
    search_box.clear()
    search_box.send_keys(keyword)
    time.sleep(random.uniform(1.5, 2.5))

    suggestions = driver.find_elements(By.XPATH, '//ul[@role="listbox"]//li')
    return [item.text for item in suggestions]

# Write results back to the Excel file
for row_idx, keyword in enumerate(data, start=3):  # Start from row 3
    if keyword:
        suggestions = fetch_suggestions(keyword)

        if suggestions:
            shortest = min(suggestions, key=len)
            longest = max(suggestions, key=len)
        else:
            shortest = None
            longest = None

        # Write results to columns D and E
        sheet[f"D{row_idx}"].value = longest
        sheet[f"E{row_idx}"].value = shortest

# Save the updated workbook
workbook.save(file_path)
print(f"Results have been written to {file_path}")

# Close the browser
driver.quit()
